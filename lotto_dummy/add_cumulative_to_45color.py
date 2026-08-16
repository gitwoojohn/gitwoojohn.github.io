"""45ColorData 구조(회차당 4열 블록)의 4번째 빈 열에 누적 비율을 채우는 모듈.

- 4번째 빈 열에 번호별 "출현 누적 %"를 기록 (보너스는 당첨번호의 일종으로 '출현'에 포함)
- 당첨/보너스 번호 셀에는 해당 색 배경
- 원본은 수정하지 않고 "{파일명}_누적.xlsx" 새 파일로 저장
"""
import os
import sys

import openpyxl
from openpyxl.styles import PatternFill

from winning_analysis_loader import DataLoader as DL


def analyze_worksheet(ws):
    """시트가 45ColorData 구조인지 검증.

    유효하면 (cols, win_color, bonus_color) 반환, 아니면 None.
    cols = [(col_idx, round_id), ...] (col_idx 0 기준)
    """
    headers, rows = DL._read_sheet(ws)
    cols = DL._round_columns(headers, rows)
    if not cols:
        return None
    win_color, bonus_color = DL._detect_colors(cols, rows)
    if win_color is None:
        return None
    return cols, win_color, bonus_color


def fill_cumulative(ws, cols, win_color, bonus_color):
    """회차 블록의 4번째 빈 열에 출현 누적 % + 배경색을 기록."""
    headers, rows = DL._read_sheet(ws)
    appeared_cnt = {n: 0 for n in range(1, 46)}
    win_fill = PatternFill(start_color=win_color, end_color=win_color, fill_type="solid")
    bonus_fill = PatternFill(start_color=bonus_color, end_color=bonus_color, fill_type="solid")
    elapsed = 0
    first = True
    for col_idx, _rid in cols:
        elapsed += 1
        num_to_row = {}
        win_r = set()
        bonus_r = set()
        for i, row in enumerate(rows):
            if col_idx >= len(row):
                continue
            v, fill = row[col_idx]
            if v is None or not isinstance(v, (int, float)):
                continue
            n = int(v)
            if 1 <= n <= 45:
                num_to_row[n] = i + 2
            if fill == win_color:
                appeared_cnt[n] += 1
                win_r.add(n)
            elif fill == bonus_color:
                appeared_cnt[n] += 1
                bonus_r.add(n)
        for n in range(1, 46):
            r = num_to_row.get(n)
            if r is None:
                continue
            cell = ws.cell(row=r, column=col_idx + 4, value=appeared_cnt[n] / elapsed)
            cell.number_format = "0.00%"
            if n in win_r:
                cell.fill = win_fill
            elif n in bonus_r:
                cell.fill = bonus_fill
        if first:
            ws.cell(row=1, column=col_idx + 4, value="출현 누적%")
            first = False


def add_cumulative(filepath):
    """파일의 첫 번째 시트에 누적 비율을 추가해 새 파일로 저장."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    result = analyze_worksheet(ws)
    if result is None:
        print("선택 시트가 45ColorData 구조가 아닙니다.")
        return False
    cols, win_color, bonus_color = result
    fill_cumulative(ws, cols, win_color, bonus_color)
    stem, ext = os.path.splitext(filepath)
    out = f"{stem}_누적{ext}"
    wb.save(out)
    print(f"완료: {out} | 회차 {len(cols)}개 처리")
    return True


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else "Test45ColorData.xlsx"
    add_cumulative(target)
