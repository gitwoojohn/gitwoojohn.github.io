import random

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from winning_analysis_constants import (
    COLOR_SIX,
    COLOR_BONUS,
    SHEET_DATA,
    SHEET_RATIO,
    SHEET_GROUP,
)

ROUNDS = 50
NUMBERS = list(range(1, 46))
GROUPS = {"그룹1(2-16)": range(2, 17), "그룹2(17-31)": range(17, 32), "그룹3(32-46)": range(32, 47)}


def create_data_sheet(ws):
    for col in range(1, ROUNDS + 1):
        ws.cell(row=1, column=col, value=col)
    for col in range(1, ROUNDS + 1):
        win_numbers = set(random.sample(NUMBERS, 6))
        bonus_numbers = set(random.sample([n for n in NUMBERS if n not in win_numbers], 1))
        shuffled = NUMBERS[:]
        random.shuffle(shuffled)
        for row, num in enumerate(shuffled, start=2):
            cell = ws.cell(row=row, column=col, value=num)
            if num in win_numbers:
                cell.fill = PatternFill(start_color=COLOR_SIX, end_color=COLOR_SIX, fill_type="solid")
            elif num in bonus_numbers:
                cell.fill = PatternFill(start_color=COLOR_BONUS, end_color=COLOR_BONUS, fill_type="solid")


def _rgb(cell):
    return (cell.fill.fgColor.rgb or "")[-6:].upper()


def create_ratio_sheet(wb, ws):
    counts = {n: 0 for n in NUMBERS}
    for col in range(1, ROUNDS + 1):
        for row in range(2, 47):
            cell = ws.cell(row=row, column=col)
            if _rgb(cell) == COLOR_SIX:
                counts[cell.value] += 1

    if SHEET_RATIO in wb.sheetnames:
        del wb[SHEET_RATIO]
    ws2 = wb.create_sheet(SHEET_RATIO)
    ws2.cell(row=1, column=1, value="번호")
    ws2.cell(row=1, column=2, value="당첨 횟수")
    ws2.cell(row=1, column=3, value="당첨 비율")
    for i, n in enumerate(NUMBERS, start=2):
        ws2.cell(row=i, column=1, value=n)
        ws2.cell(row=i, column=2, value=counts[n])
        ws2.cell(row=i, column=3, value=f"{counts[n] / ROUNDS:.2%}")


def create_group_sheet(wb, ws):
    if SHEET_GROUP in wb.sheetnames:
        del wb[SHEET_GROUP]
    ws2 = wb.create_sheet(SHEET_GROUP)

    ws2.cell(row=1, column=1, value="회차")
    for col, name in enumerate(GROUPS, start=2):
        ws2.cell(row=1, column=col, value=name)

    for round_num in range(1, ROUNDS + 1):
        ws2.cell(row=round_num + 1, column=1, value=round_num)
        for col, name in enumerate(GROUPS, start=2):
            count = sum(
                1
                for row in GROUPS[name]
                if _rgb(ws.cell(row=row, column=round_num)) in (COLOR_SIX, COLOR_BONUS)
            )
            ws2.cell(row=round_num + 1, column=col, value=count)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_DATA
    create_data_sheet(ws)
    create_ratio_sheet(wb, ws)
    create_group_sheet(wb, ws)
    wb.save("lotto_dummy.xlsx")
    print("done")


if __name__ == "__main__":
    main()
