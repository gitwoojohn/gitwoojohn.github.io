from PySide6.QtCore import QThread, Signal
from openpyxl import load_workbook

from winning_analysis_constants import SHEET_DATA


def cell_fill_rgb(cell):
    try:
        rgb = cell.fill.fgColor.rgb
    except Exception:
        return ""
    if rgb is None:
        return ""
    return str(rgb).upper()


def compute_stats(numbers, win, bonus):
    win_sorted = sorted(win)
    if not win_sorted:
        return {
            "total": 0, "ac": 0, "sd": 0, "odd": 0, "even": 0,
            "group_counts": [0, 0, 0],
        }
    total = sum(win_sorted)
    diffs = {
        abs(win_sorted[j] - win_sorted[i])
        for i in range(len(win_sorted))
        for j in range(i + 1, len(win_sorted))
    }
    ac = len(diffs) - (len(win_sorted) - 1)
    mean = total / len(win_sorted)
    sd = (sum((n - mean) ** 2 for n in win_sorted) / len(win_sorted)) ** 0.5
    odd = sum(1 for n in win_sorted if n % 2 == 1)
    even = len(win_sorted) - odd
    marked = win | bonus
    group_counts = [0, 0, 0]
    for idx, n in enumerate(numbers):
        if n in marked:
            g = idx // 15
            if 0 <= g < 3:
                group_counts[g] += 1
    return {
        "total": total,
        "ac": ac,
        "sd": sd,
        "odd": odd,
        "even": even,
        "group_counts": group_counts,
    }


class SheetLister(QThread):
    listed = Signal(list)
    failed = Signal(str)

    def __init__(self, filepath, parent=None):
        super().__init__(parent)
        self.filepath = filepath

    def run(self):
        try:
            wb = load_workbook(self.filepath, data_only=True, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            self.listed.emit(sheets)
        except Exception as e:
            self.failed.emit(f"파일을 읽을 수 없습니다: {e}")


class DataLoader(QThread):
    loaded = Signal(dict)
    failed = Signal(str)

    def __init__(self, filepath, sheet_name=SHEET_DATA, parent=None):
        super().__init__(parent)
        self.filepath = filepath
        self.sheet_name = sheet_name

    @staticmethod
    def _read_sheet(ws):
        headers = []
        rows = []
        for r, row in enumerate(ws.iter_rows()):
            if r == 0:
                headers = [c.value for c in row]
            else:
                rows.append([(c.value, cell_fill_rgb(c)[-6:]) for c in row])
        return headers, rows

    @staticmethod
    def _round_columns(headers, rows):
        cols = []
        for c, header in enumerate(headers):
            if isinstance(header, (int, float)) and not isinstance(header, bool):
                cnt = sum(
                    1 for row in rows
                    if c < len(row) and isinstance(row[c][0], (int, float))
                )
                if cnt >= 40:
                    cols.append((c, int(header)))
        return cols

    @classmethod
    def _is_data_sheet(cls, ws):
        headers, rows = cls._read_sheet(ws)
        return len(cls._round_columns(headers, rows)) > 0, headers, rows

    @staticmethod
    def _detect_colors(cols, rows):
        from collections import Counter
        counter = Counter()
        for c, _ in cols:
            for row in rows:
                if c >= len(row):
                    continue
                fill = row[c][1]
                if fill and fill != "000000":
                    counter[fill] += 1
        top = counter.most_common(2)
        win_color = top[0][0] if top else None
        bonus_color = top[1][0] if len(top) > 1 else None
        return win_color, bonus_color

    def run(self):
        try:
            wb = load_workbook(self.filepath, data_only=True, read_only=True)
        except Exception as e:
            self.failed.emit(f"파일을 읽을 수 없습니다: {e}")
            return
        try:
            grids = {}
            data_sheets = []
            for sn in wb.sheetnames:
                is_data, headers, rows = self._is_data_sheet(wb[sn])
                grids[sn] = (headers, rows)
                if is_data:
                    data_sheets.append(sn)
            if not data_sheets:
                data_sheets = wb.sheetnames
            if self.sheet_name in data_sheets:
                target = self.sheet_name
            else:
                target = data_sheets[0]
            headers, rows = grids[target]
            wb.close()

            cols = self._round_columns(headers, rows)
            if not cols:
                raise ValueError("회차 열을 찾을 수 없습니다 (첫 행에 회차 번호 필요)")
            round_ids = [r for _, r in cols]
            rounds = len(round_ids)
            win_color, bonus_color = self._detect_colors(cols, rows)

            numbers_by_round = {}
            win_by_round = {}
            bonus_by_round = {}
            for col_idx, round_id in cols:
                numbers = []
                win = set()
                bonus = set()
                for row in rows:
                    if col_idx >= len(row):
                        continue
                    v, fill = row[col_idx]
                    if v is None:
                        continue
                    numbers.append(v)
                    if fill == win_color:
                        win.add(v)
                    elif fill == bonus_color:
                        bonus.add(v)
                numbers_by_round[round_id] = numbers
                win_by_round[round_id] = win
                bonus_by_round[round_id] = bonus

            ratio = {n: (0, "0.00%") for n in range(1, 46)}
            for round_id in round_ids:
                for n in win_by_round[round_id]:
                    if n is None:
                        continue
                    cnt, _ = ratio[n]
                    ratio[n] = (cnt + 1, f"{(cnt + 1) / rounds:.2%}")

            group = []
            group_headers = ["회차", "그룹1(1-15)", "그룹2(16-30)", "그룹3(31-45)"]
            for round_id in round_ids:
                marked = win_by_round[round_id] | bonus_by_round[round_id]
                counts = [0, 0, 0]
                for n in marked:
                    if n is None:
                        continue
                    g = (n - 1) // 15
                    if 0 <= g < 3:
                        counts[g] += 1
                group.append([round_id] + counts)
        except Exception as e:
            self.failed.emit(f"데이터를 분석하는 중 오류가 발생했습니다: {e}")
            return

        self.loaded.emit({
            "rounds": rounds,
            "round_ids": round_ids,
            "numbers": numbers_by_round,
            "win": win_by_round,
            "bonus": bonus_by_round,
            "ratio": ratio,
            "group_headers": group_headers,
            "group": group,
            "sheets": data_sheets,
            "sheet_name": target,
        })
