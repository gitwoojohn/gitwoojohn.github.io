import os
from datetime import datetime

from PySide6.QtCore import Qt
from PySide6.QtGui import QBrush, QColor
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QHeaderView, QLabel, QPushButton, QApplication,
)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

_SAVE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


class NumericTableItem(QTableWidgetItem):
    def __init__(self, text, value):
        super().__init__(text)
        self.setData(Qt.UserRole, value)

    def __lt__(self, other):
        if isinstance(other, NumericTableItem):
            return self.data(Qt.UserRole) < other.data(Qt.UserRole)
        return super().__lt__(other)


class RatioTab(QWidget):
    def __init__(self, data, round_id=None, parent=None):
        super().__init__(parent)
        self.data = data
        self.round_ids = data.get("round_ids", [])
        self.last_round = self.round_ids[-1] if self.round_ids else 0
        self.current_round_id = None
        self.current_counts = {}
        self.current_used = 0

        layout = QVBoxLayout(self)

        top = QHBoxLayout()
        self.rangeLabel = QLabel("", self)
        self.rangeLabel.setAlignment(Qt.AlignCenter)
        self.rangeLabel.setObjectName("loading")
        top.addWidget(self.rangeLabel, 1)
        self.copyButton = QPushButton("복사", self)
        self.copyButton.setFixedWidth(60)
        self.copyButton.clicked.connect(self._copy)
        top.addWidget(self.copyButton)
        self.saveButton = QPushButton("엑셀 저장", self)
        self.saveButton.setFixedWidth(90)
        self.saveButton.clicked.connect(self._save_all)
        top.addWidget(self.saveButton)
        layout.addLayout(top)

        table = QTableWidget(45, 4, self)
        table.setHorizontalHeaderLabels(["번호", "누적 당첨 횟수", "누적 당첨 비율", "다음 회차 출현"])
        table.setAlternatingRowColors(True)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self._num_items = {}
        for i, num in enumerate(range(1, 46)):
            it0 = NumericTableItem(str(num), num)
            it1 = NumericTableItem("0", 0)
            it2 = NumericTableItem("0.00%", 0.0)
            it3 = QTableWidgetItem("미출현")
            for it in (it0, it1, it2, it3):
                it.setTextAlignment(Qt.AlignCenter)
            table.setItem(i, 0, it0)
            table.setItem(i, 1, it1)
            table.setItem(i, 2, it2)
            table.setItem(i, 3, it3)
            self._num_items[num] = (it0, it1, it2, it3)
        table.setSortingEnabled(True)
        layout.addWidget(table)

        self.table = table
        self.update_for_round(round_id if round_id is not None else self.last_round)

    def _next_round(self, round_id):
        try:
            idx = self.round_ids.index(round_id)
        except ValueError:
            return None
        if idx + 1 < len(self.round_ids):
            return self.round_ids[idx + 1]
        return None

    def update_for_round(self, round_id):
        prev_rounds = [rid for rid in self.round_ids if rid <= round_id]
        used = len(prev_rounds)
        counts = {n: 0 for n in range(1, 46)}
        for rid in prev_rounds:
            for n in self.data["win"].get(rid, ()):
                if n is not None and 1 <= n <= 45:
                    counts[n] += 1
        self.current_round_id = round_id
        self.current_counts = counts
        self.current_used = used
        next_rid = self._next_round(round_id)
        win_next = self.data["win"].get(next_rid, set()) if next_rid is not None else set()
        bonus_next = self.data["bonus"].get(next_rid, set()) if next_rid is not None else set()
        used_next = used + 1
        for num in range(1, 46):
            it0, it1, it2, it3 = self._num_items[num]
            pct = counts[num] / used if used else 0.0
            it1.setText(str(counts[num]))
            it1.setData(Qt.UserRole, counts[num])
            it2.setText(f"{pct:.2%}")
            it2.setData(Qt.UserRole, pct)
            if num in win_next:
                new_pct = (counts[num] + 1) / used_next if used_next else 0.0
                it3.setText(f"당첨 ({new_pct:.2%})")
                it3.setBackground(QColor("#00c0ff"))
                it3.setForeground(QColor("#000000"))
            elif num in bonus_next:
                new_pct = (counts[num] + 1) / used_next if used_next else 0.0
                it3.setText(f"보너스 ({new_pct:.2%})")
                it3.setBackground(QColor("#ffbb00"))
                it3.setForeground(QColor("#000000"))
            else:
                it3.setText("미출현")
                it3.setBackground(QBrush())
                it3.setForeground(QBrush())
        if next_rid is not None:
            self.table.horizontalHeaderItem(3).setText(f"다음 회차 출현 ({next_rid}회차)")
        else:
            self.table.horizontalHeaderItem(3).setText("다음 회차 출현")
        start = self.round_ids[0] if self.round_ids else round_id
        if next_rid is not None:
            self.rangeLabel.setText(f"누적 {start}회차 ~ {round_id}회차 | 다음 회차: {next_rid}회차")
        else:
            self.rangeLabel.setText(f"누적 {start}회차 ~ {round_id}회차 (마지막 회차)")

    def _copy_text(self):
        headers = [self.table.horizontalHeaderItem(c).text() for c in range(self.table.columnCount())]
        lines = ["\t".join(headers)]
        for r in range(self.table.rowCount()):
            lines.append("\t".join(self.table.item(r, c).text() for c in range(self.table.columnCount())))
        return self.rangeLabel.text() + "\n" + "\n".join(lines)

    def _copy(self):
        QApplication.clipboard().setText(self._copy_text())

    def _save_all(self):
        path = os.path.join(
            _SAVE_DIR,
            f"누적비율_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        self._save_excel(path)
        self._verify_excel(path)

    def _save_excel(self, path):
        if not self.round_ids:
            return
        cum = {n: 0 for n in range(1, 46)}
        cum_per_round = []
        for rid in self.round_ids:
            for n in self.data["win"].get(rid, ()):
                if n is not None and 1 <= n <= 45:
                    cum[n] += 1
            cum_per_round.append(dict(cum))

        win_fill = PatternFill(start_color="00C0FF", end_color="00C0FF", fill_type="solid")
        bonus_fill = PatternFill(start_color="FFBB00", end_color="FFBB00", fill_type="solid")

        if os.path.exists(path):
            wb = load_workbook(path)
            if "누적횟수" in wb.sheetnames:
                ws_cnt = wb["누적횟수"]
            else:
                ws_cnt = wb.active
                ws_cnt.title = "누적횟수"
            if "누적비율" in wb.sheetnames:
                ws_ratio = wb["누적비율"]
            else:
                ws_ratio = wb.create_sheet("누적비율")
                ws_ratio.cell(row=1, column=1, value="번호")
                for num in range(1, 46):
                    ws_ratio.cell(row=num + 1, column=1, value=num)
            all_cols = {}
            for c in range(2, ws_cnt.max_column + 1):
                v = ws_cnt.cell(row=1, column=c).value
                if v is not None:
                    all_cols[int(v)] = c
        else:
            wb = Workbook()
            ws_cnt = wb.active
            ws_cnt.title = "누적횟수"
            ws_ratio = wb.create_sheet("누적비율")
            for ws in (ws_cnt, ws_ratio):
                ws.cell(row=1, column=1, value="번호")
                for num in range(1, 46):
                    ws.cell(row=num + 1, column=1, value=num)
            all_cols = {}

        for rid in self.round_ids:
            if rid in all_cols:
                continue
            idx = self.round_ids.index(rid)
            col = ws_cnt.max_column + 1
            ws_cnt.cell(row=1, column=col, value=rid)
            ws_ratio.cell(row=1, column=col, value=rid)
            used = idx + 1
            for num in range(1, 46):
                cnt = cum_per_round[idx][num]
                ws_cnt.cell(row=num + 1, column=col, value=cnt)
                ws_ratio.cell(row=num + 1, column=col, value=round(cnt / used, 6) if used else 0.0)
            all_cols[rid] = col

        for rid, col in all_cols.items():
            win = self.data["win"].get(rid, set())
            bonus = self.data["bonus"].get(rid, set())
            for num in range(1, 46):
                ratio_cell = ws_ratio.cell(row=num + 1, column=col)
                ratio_cell.number_format = "0.00%"
                if num in win:
                    ws_cnt.cell(row=num + 1, column=col).fill = win_fill
                    ratio_cell.fill = win_fill
                elif num in bonus:
                    ws_cnt.cell(row=num + 1, column=col).fill = bonus_fill
                    ratio_cell.fill = bonus_fill
        wb.save(path)

    def _verify_excel(self, path):
        if not self.round_ids:
            return
        windows = [100, 50, 30, 10]
        buckets = [(i, i + 1) for i in range(0, 30)] + [(30.0, 101.0)]
        pre_cum = {}
        cum = {n: 0 for n in range(1, 46)}
        seen = 0
        for rid in self.round_ids:
            pre_cum[rid] = (dict(cum), seen)
            for n in self.data["win"].get(rid, ()):
                if n is not None and 1 <= n <= 45:
                    cum[n] += 1
            seen += 1

        if os.path.exists(path):
            wb = load_workbook(path)
        else:
            wb = Workbook()
            wb.remove(wb.active)
        for win in windows:
            sheet_name = f"최근{win}회"
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            recent = self.round_ids[-win:]
            ws = wb.create_sheet(sheet_name)
            ws.append(["누적 비율 구간(%)", "번호 건수", "다음 회차 당첨 횟수", "당첨률(%)"])
            bucket_stats = {b: [0, 0] for b in buckets}
            for rid in recent:
                pre, idx = pre_cum[rid]
                if idx == 0:
                    continue
                win_r = self.data["win"].get(rid, set())
                for num in range(1, 46):
                    ratio = pre[num] / idx * 100
                    bucket = None
                    for lo, hi in buckets:
                        if lo <= ratio < hi:
                            bucket = (lo, hi)
                            break
                    if bucket is None:
                        continue
                    bucket_stats[bucket][0] += 1
                    if num in win_r:
                        bucket_stats[bucket][1] += 1
            for (lo, hi), (cnt, hits) in bucket_stats.items():
                rate = hits / cnt * 100 if cnt else 0.0
                ws.append([f"{lo:.0f}~{hi:.0f}", cnt, hits, round(rate, 2)])
        wb.save(path)
