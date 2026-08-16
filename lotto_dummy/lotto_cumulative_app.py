"""45ColorData 누적 비율 추가 앱 (PySide6, 다크테마).

- 엑셀 파일 선택, 시트 선택
- 선택 시트가 45ColorData 구조(회차 열/당첨·보너스 색상)인지 검증
- 적합하면 4번째 빈 열에 출현 누적 %를 추가해 새 파일로 저장
- 적합하지 않으면 메시지 표시
"""
import os
import sys

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout,
    QLabel, QPushButton, QComboBox, QFileDialog, QMessageBox,
)
from openpyxl import load_workbook

from winning_analysis_loader import SheetLister
from winning_analysis_styles import build_arrows, build_qss
from add_cumulative_to_45color import analyze_worksheet, fill_cumulative

APP_DIR = os.path.dirname(os.path.abspath(__file__))


class CumulativeWindow(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("45ColorData 누적 비율 추가")
        self.resize(720, 150)

        central = QWidget(self)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(8)

        top = QHBoxLayout()
        self.openButton = QPushButton("파일 불러오기", central)
        self.openButton.clicked.connect(self.on_open_clicked)
        self.fileLabel = QLabel("파일: -", central)
        self.fileLabel.setObjectName("fileInfo")
        self.sheetCombo = QComboBox(central)
        self.sheetCombo.setMinimumWidth(200)
        self.sheetCombo.setEnabled(False)
        self.sheetCombo.currentIndexChanged.connect(self.on_sheet_changed)
        top.addWidget(self.openButton)
        top.addSpacing(8)
        top.addWidget(self.fileLabel)
        top.addSpacing(8)
        top.addWidget(QLabel("시트:", central))
        top.addWidget(self.sheetCombo)
        top.addStretch(1)
        layout.addLayout(top)

        self.statusLabel = QLabel("엑셀 파일을 선택하세요.", central)
        self.statusLabel.setObjectName("loading")
        self.statusLabel.setWordWrap(True)
        layout.addWidget(self.statusLabel)

        bottom = QHBoxLayout()
        bottom.addStretch(1)
        self.saveButton = QPushButton("누적% 추가 저장", central)
        self.saveButton.setEnabled(False)
        self.saveButton.clicked.connect(self.on_save_clicked)
        bottom.addWidget(self.saveButton)
        layout.addLayout(bottom)

        self.setCentralWidget(central)

        self.filepath = None
        self.sheet_lister = None
        self.valid = False

    def on_open_clicked(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "엑셀 파일 선택", APP_DIR, "Excel Files (*.xlsx *.xlsm)"
        )
        if not path:
            return
        self.filepath = path
        self.fileLabel.setText(f"파일: {os.path.basename(path)}")
        self.sheetCombo.blockSignals(True)
        self.sheetCombo.clear()
        self.sheetCombo.blockSignals(False)
        self.sheetCombo.setEnabled(False)
        self.saveButton.setEnabled(False)
        self.valid = False
        self.statusLabel.setText("시트 목록 확인 중...")
        self.sheet_lister = SheetLister(path, self)
        self.sheet_lister.listed.connect(self.on_sheets_listed)
        self.sheet_lister.failed.connect(self.on_load_failed)
        self.sheet_lister.start()

    def on_sheets_listed(self, sheets):
        self.sheetCombo.blockSignals(True)
        self.sheetCombo.clear()
        self.sheetCombo.addItems(sheets)
        self.sheetCombo.setCurrentIndex(-1)
        self.sheetCombo.blockSignals(False)
        self.sheetCombo.setEnabled(True)
        self.statusLabel.setText("시트를 선택하세요.")

    def on_load_failed(self, message):
        QMessageBox.warning(self, "로드 실패", message)

    def on_sheet_changed(self, index):
        if index < 0:
            return
        self.validate_current()

    def validate_current(self):
        if not self.filepath:
            return
        sheet = self.sheetCombo.currentText()
        if not sheet:
            return
        try:
            wb = load_workbook(self.filepath, data_only=False)
            ws = wb[sheet]
            result = analyze_worksheet(ws)
            wb.close()
        except Exception as e:
            self.valid = False
            self.saveButton.setEnabled(False)
            self.statusLabel.setText(f"시트 읽기 오류: {e}")
            return
        if result is None:
            self.valid = False
            self.saveButton.setEnabled(False)
            self.statusLabel.setText("선택 시트는 45ColorData 구조(회차 열/색상)가 아닙니다.")
        else:
            self.valid = True
            self.saveButton.setEnabled(True)
            self.statusLabel.setText("선택 시트 확인 완료 (회차 열/색상 감지됨) — 저장 가능")

    def on_save_clicked(self):
        if not self.filepath:
            QMessageBox.warning(self, "알림", "엑셀 파일을 선택하세요.")
            return
        sheet = self.sheetCombo.currentText()
        if not sheet:
            QMessageBox.warning(self, "알림", "시트를 선택하세요.")
            return
        try:
            wb = load_workbook(self.filepath, data_only=False)
            ws = wb[sheet]
            result = analyze_worksheet(ws)
            if result is None:
                wb.close()
                QMessageBox.warning(self, "시트 확인", "선택한 시트는 45ColorData 구조가 아닙니다.")
                return
            cols, win_color, bonus_color = result
            fill_cumulative(ws, cols, win_color, bonus_color)
            stem, ext = os.path.splitext(self.filepath)
            out = f"{stem}_누적{ext}"
            wb.save(out)
            wb.close()
        except Exception as e:
            QMessageBox.warning(self, "오류", f"처리 중 오류가 발생했습니다: {e}")
            return
        self.statusLabel.setText(f"저장 완료: {os.path.basename(out)}")
        QMessageBox.information(self, "완료", f"저장되었습니다:\n{out}")


def main():
    app = QApplication(sys.argv)
    build_arrows()
    app.setStyleSheet(build_qss())
    win = CumulativeWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
