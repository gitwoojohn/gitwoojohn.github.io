import pandas as pd
import numpy as np
from openpyxl.styles import Alignment, PatternFill, Font

class LottoGapAnalyzer:
    def __init__(self, filepath):
        self.df = pd.read_csv(filepath)
        
        # 1207회차 데이터가 없다면 수동 추가
        if self.df['회차'].max() < 1207:
            new_row = {
                '회차': 1207,
                '1번': 10, '2번': 22, '3번': 24, '4번': 27, '5번': 38, '6번': 45, 
                '보너스': 11, '합계': 166
            }
            self.df = pd.concat([self.df, pd.DataFrame([new_row])], ignore_index=True)
            
        self.df = self.df.sort_values('회차').reset_index(drop=True)

    def run_analysis(self):
        print("--- 번호별 출현 간격(이격) 정밀 추적 중 ---")
        
        # 분석 결과를 담을 리스트
        gap_results = []
        
        # 번호별 마지막 등장 회차를 기록할 딕셔너리 (초기값 -1)
        # {1: 1200, 2: 1198 ...}
        last_seen = {i: -1 for i in range(1, 46)}
        
        cols = ['1번', '2번', '3번', '4번', '5번', '6번']

        for idx, row in self.df.iterrows():
            current_round = row['회차']
            current_nums = row[cols].values
            
            # 이번 회차 번호들의 '이격(Gap)' 계산
            # 이격 = 현재회차 - 직전등장회차 - 1 (바로 연달아 나오면 이격 0)
            round_gaps = []
            
            for num in current_nums:
                last_r = last_seen[num]
                if last_r == -1:
                    gap = 0 # 처음에 등장할 땐 0으로 간주 (데이터 부족)
                else:
                    gap = current_round - last_r - 1
                round_gaps.append(gap)
                
                # '마지막 등장 회차' 업데이트
                last_seen[num] = current_round
                
            # 통계 도출
            avg_gap = np.mean(round_gaps)
            max_gap = np.max(round_gaps) # 가장 오래 묵은 번호의 이격
            
            # 핫/콜드 분류
            # Hot(0~4회 쉼), Warm(5~9회 쉼), Cold(10회 이상 쉼)
            hot_cnt = sum(1 for g in round_gaps if g < 5)
            cold_cnt = sum(1 for g in round_gaps if g >= 10)
            
            gap_results.append({
                '회차': current_round,
                '번호1_이격': round_gaps[0],
                '번호2_이격': round_gaps[1],
                '번호3_이격': round_gaps[2],
                '번호4_이격': round_gaps[3],
                '번호5_이격': round_gaps[4],
                '번호6_이격': round_gaps[5],
                '평균이격': round(avg_gap, 1),
                '최대이격(묵은지)': max_gap,
                '핫(Hot)개수': hot_cnt,
                '콜드(Cold)개수': cold_cnt
            })
            
        # 데이터프레임 변환
        result_df = pd.DataFrame(gap_results)
        
        # 엑셀 저장
        file_name = "로또_이격분석_인사이트.xlsx"
        
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name='이격분석', index=False)
            
            # 스타일링 (조건부 서식 흉내)
            ws = writer.sheets['이격분석']
            ws.freeze_panes = 'A2'
            ws.row_dimensions[1].height = 50
            
            center_align = Alignment(horizontal='center', vertical='center')
            
            # 색상 정의
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid') # Cold (위험/주목)
            blue_fill = PatternFill(start_color='E0E0FF', end_color='E0E0FF', fill_type='solid') # Hot (평범)
            
            # Gap 컬럼 인덱스 찾기
            gap_col_indices = []
            for cell in ws[1]:
                if '이격' in str(cell.value):
                    gap_col_indices.append(cell.column)
            
            for col in ws.columns:
                # 너비 자동 조절 (Auto Fit)
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        val = str(cell.value)
                        # 실수는 소수점 정리해서 길이 측정
                        if isinstance(cell.value, float): val = f"{cell.value:.1f}"
                        if len(val) > max_len: max_len = len(val)
                    except: pass
                ws.column_dimensions[col_letter].width = max((max_len + 2) * 1.3, 8)
                
                for cell in col:
                    cell.alignment = center_align
                    
                    # 헤더 제외하고 내용만
                    if cell.row > 1:
                        # 이격 데이터 하이라이팅
                        if cell.column in gap_col_indices:
                            try:
                                val = float(cell.value)
                                if val >= 10: # 10회 이상 쉬었다 나온 경우 (Cold)
                                    cell.fill = red_fill
                                    cell.font = Font(bold=True)
                                elif val < 5: # 자주 나온 경우 (Hot)
                                    cell.fill = blue_fill
                            except: pass

        print(f"💾 엑셀 저장 완료: {file_name}")
        print(f"   - 🔴 빨간색 셀: 10회 이상 숨어있다 튀어나온 '대박 이격'")
        print(f"   - 🔵 파란색 셀: 5회 미만으로 자주 나오는 '단골 이격'")

if __name__ == "__main__":
    analyzer = LottoGapAnalyzer('ReadWeb-WinningNumbers.csv')
    analyzer.run_analysis()
