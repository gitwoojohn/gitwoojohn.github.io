import pandas as pd
import numpy as np
from openpyxl.styles import Alignment, Font, PatternFill

class LottoRhythmFinder:
    def __init__(self, filepath):
        self.df = pd.read_csv(filepath)
        
        # 1207회차 자동 추가
        if self.df['회차'].max() < 1207:
            new_row = {
                '회차': 1207,
                '1번': 10, '2번': 22, '3번': 24, '4번': 27, '5번': 38, '6번': 45, '보너스': 11
            }
            self.df = pd.concat([self.df, pd.DataFrame([new_row])], ignore_index=True)
            
        self.current_round = self.df['회차'].max()

    def run_analysis(self):
        print(f"--- [제{self.current_round}회 기준] 번호별 회귀 본능(리듬) 분석 중 ---")
        
        # 1. 모든 번호의 출현 회차 수집
        num_appearances = {i: [] for i in range(1, 46)}
        cols = ['1번', '2번', '3번', '4번', '5번', '6번', '보너스'] # 보너스 포함 흐름 파악
        
        for idx, row in self.df.iterrows():
            r_num = row['회차']
            for col in cols:
                num_appearances[row[col]].append(r_num)
                
        # 2. 주기 및 규칙성 계산
        results = []
        
        for num in range(1, 46):
            apps = sorted(num_appearances[num])
            
            if len(apps) < 2:
                continue # 데이터 부족
            
            # 간격(Cycle) 리스트 계산: [10회, 12회, 11회 ...]
            cycles = np.diff(apps)
            
            # 현재 마지막 등장 이후 흐른 시간 (Current Gap)
            last_seen = apps[-1]
            current_wait = self.current_round - last_seen
            
            # 통계 도출
            avg_cycle = np.mean(cycles)      # 평균 복귀 주기
            std_dev = np.std(cycles)         # 규칙성 (낮을수록 칼같이 지킴)
            
            # 타이밍 판단 (Z-Score 개념 약식 적용)
            # (현재대기 - 평균주기)
            diff_from_avg = current_wait - avg_cycle
            
            status = ""
            # 규칙적인 놈(표준편차 10 이하)이 평균 주기에 근접했을 때
            if std_dev < 12: 
                if -3 <= diff_from_avg <= 3:
                    status = "⏰ 도착임박 (D-Day)"
                elif diff_from_avg > 3:
                    status = "🚀 지연도착 (과적)"
                elif diff_from_avg < -3:
                    status = "💤 휴식중"
            else:
                status = "🎲 불규칙함"

            results.append({
                '번호': num,
                '총출현': len(apps),
                '평균주기': round(avg_cycle, 1),
                '규칙성(낮을수록좋음)': round(std_dev, 1),
                '현재대기(Gap)': current_wait,
                '예측차이': round(diff_from_avg, 1),
                '상태': status
            })
            
        # 데이터프레임 변환 및 정렬 (규칙성 좋은 순서)
        result_df = pd.DataFrame(results)
        result_df = result_df.sort_values('규칙성(낮을수록좋음)')
        
        # 3. 엑셀 저장
        file_name = f"로또_규칙적회귀분석_{self.current_round}회.xlsx"
        
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name='전체랭킹', index=False)
            
            # [핵심] 도착 임박한 모범생들만 따로 모음
            target_mask = result_df['상태'].str.contains('도착|지연')
            target_df = result_df[target_mask].sort_values('규칙성(낮을수록좋음)')
            target_df.to_excel(writer, sheet_name='추천_타겟번호', index=False)
            
            # 스타일링
            center_align = Alignment(horizontal='center', vertical='center')
            
            # 도착임박(노랑), 지연(빨강)
            yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

            for sheet_name in writer.sheets:
                ws = writer.sheets[sheet_name]
                ws.freeze_panes = 'A2'
                ws.row_dimensions[1].height = 40
                
                # 상태 컬럼 찾기
                status_col = None
                for cell in ws[1]:
                    if cell.value == '상태': status_col = cell.column

                # AutoFit 및 정렬
                for col in ws.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            val_str = str(cell.value)
                            if len(val_str) > max_len: max_len = len(val_str)
                        except: pass
                    ws.column_dimensions[col_letter].width = max((max_len + 2) * 1.3, 8)
                    
                    for cell in col:
                        cell.alignment = center_align
                        # 조건부 서식 적용
                        if cell.column == status_col and cell.row > 1:
                            if '도착' in str(cell.value):
                                cell.fill = yellow_fill
                                cell.font = Font(bold=True)
                            elif '지연' in str(cell.value):
                                cell.fill = red_fill
                                cell.font = Font(bold=True)

        print(f"💾 엑셀 저장 완료: {file_name}")
        print(f"   - Sheet2 [추천_타겟번호]를 꼭 확인하세요.")
        print(f"   - '도착임박': 평균 주기에 딱 맞춰서 온 모범생 번호들입니다.")

if __name__ == "__main__":
    finder = LottoRhythmFinder('ReadWeb-WinningNumbers.csv')
    finder.run_analysis()
